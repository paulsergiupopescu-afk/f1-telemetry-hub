#!/usr/bin/env python3
"""Generate a polished lap-by-lap Excel report from two telemetry CSV files.

The workbook includes a summary, one lap sheet per player, a formula-driven
comparison, raw traces, and native speed and time-delta charts. Fastest laps are
highlighted and invalid laps are marked.
"""

import argparse
import csv
import sys

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except Exception:
        pass

try:
    import numpy as np
except ImportError:
    sys.exit("numpy is missing. Run: pip install numpy openpyxl")

# Reuse the lap logic from f1_compare as the single source of truth.
try:
    from f1_compare import extract_laps, clean_curve, pick_fastest, lap_sectors, fmt_ms
except ImportError:
    sys.exit("f1_report.py requires f1_compare.py in the same folder.")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is missing. Run: pip install openpyxl")


# ----------------------------------------------------------------------------
def _f(x, d=0.0):
    try:
        return float(x)
    except (ValueError, TypeError):
        return d


def load_rich(path):
    """Load comparison samples while retaining tyre and fuel fields."""
    out = []
    with open(path, newline="", encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            if not r.get("lap_num") or not r.get("lap_distance") or not r.get("cur_lap_ms"):
                continue
            out.append({
                "lap": int(_f(r["lap_num"])), "dist": _f(r["lap_distance"]),
                "t": _f(r["cur_lap_ms"]), "last_lap": _f(r.get("last_lap_ms")),
                "speed": _f(r.get("speed_kmh")), "thr": _f(r.get("throttle")),
                "brk": _f(r.get("brake")), "s1": _f(r.get("s1_ms")),
                "s2": _f(r.get("s2_ms")), "invalid": int(_f(r.get("lap_invalid"))),
                "tyre": (r.get("tyre_visual") or "").strip(),
                "tyre_age": r.get("tyre_age_laps") or "",
                "fuel": _f(r.get("fuel_kg")),
                "ers_dep": _f(r.get("ers_deployed_lap_j")),
                "wear": (_f(r.get("tyreWear_RL")) + _f(r.get("tyreWear_RR")) +
                         _f(r.get("tyreWear_FL")) + _f(r.get("tyreWear_FR"))) / 4.0,
            })
    if not out:
        sys.exit(f"No valid samples in {path}.")
    return out


def lap_table(samples):
    """Return lap dictionaries ready for Excel output."""
    laps = extract_laps(samples)
    rows = []
    for L in sorted(laps):
        d = laps[L]
        # Skip partial laps and laps without an official time.
        if not d["t"] or d.get("coverage", 0) < 0.5:
            continue
        s1, s2, s3 = lap_sectors(d, d["t"])
        speeds = [x["speed"] for x in d["samples"] if x["speed"] > 0]
        tyres = [x["tyre"] for x in d["samples"] if x["tyre"]]
        fuels = [x["fuel"] for x in d["samples"] if x["fuel"] > 0]
        ers = [x["ers_dep"] for x in d["samples"]]
        wears = [x["wear"] for x in d["samples"]]
        rows.append({
            "lap": L, "t_ms": d["t"], "s1": s1, "s2": s2, "s3": s3,
            "top": max(speeds) if speeds else 0,
            "avg": sum(speeds) / len(speeds) if speeds else 0,
            "valid": d["valid"],
            "tyre": max(set(tyres), key=tyres.count) if tyres else "",
            "fuel_end": fuels[-1] if fuels else 0,
            "fuel_used": (fuels[0] - fuels[-1]) if len(fuels) > 1 else 0,
            "ers_mj": (max(ers) / 1e6) if ers else 0,
            "wear_end": wears[-1] if wears else 0,
        })
    return rows, laps


# ----------------------------------------------------------------------------
# Stiluri
FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F2A44")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F2A44")
BASE_FONT = Font(name=FONT, size=10)
FAST_FILL = PatternFill("solid", fgColor="C6EFCE")      # Green: fastest lap.
INVALID_FONT = Font(name=FONT, size=10, color="C00000", italic=True)
P1_FILL = PatternFill("solid", fgColor="D9EEF5")        # cyan deschis
P2_FILL = PatternFill("solid", fgColor="F7D9EA")        # roz deschis
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def write_lap_sheet(ws, title, rows, accent_fill):
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    headers = ["Lap", "Time", "Time (s)", "S1", "S2", "S3",
               "Top (km/h)", "Avg (km/h)", "Tyres", "Wear (%)",
               "Fuel used (kg)", "Fuel remaining (kg)", "ERS (MJ)", "Valid"]
    hr = 3
    for i, h in enumerate(headers, 1):
        ws.cell(row=hr, column=i, value=h)
    style_header(ws, hr, len(headers))

    fastest_ms = min((r["t_ms"] for r in rows if r["valid"]), default=None)
    for j, r in enumerate(rows):
        rr = hr + 1 + j
        vals = [r["lap"], fmt_ms(r["t_ms"]), round(r["t_ms"] / 1000, 3),
                fmt_ms(r["s1"]), fmt_ms(r["s2"]), fmt_ms(r["s3"]),
                round(r["top"]), round(r["avg"], 1),
                r["tyre"], round(r["wear_end"], 1),
                round(r["fuel_used"], 2), round(r["fuel_end"], 2),
                round(r["ers_mj"], 2), "Yes" if r["valid"] else "No"]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=rr, column=i, value=v)
            c.font = BASE_FONT
            c.border = BORDER
            c.alignment = Alignment(horizontal="center")
        if not r["valid"]:
            for i in range(1, len(headers) + 1):
                ws.cell(row=rr, column=i).font = INVALID_FONT
        if fastest_ms is not None and r["t_ms"] == fastest_ms and r["valid"]:
            for i in range(1, len(headers) + 1):
                ws.cell(row=rr, column=i).fill = FAST_FILL

    last = hr + len(rows)
    if rows:
        # Time color scale: fastest green through slowest red.
        rng = f"C{hr+1}:C{last}"
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B"))
    widths = [6, 11, 10, 10, 10, 10, 11, 11, 12, 10, 15, 14, 9, 7]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{hr+1}"


def write_summary(ws, rows1, rows2, la, lb, n1, n2):
    ws.cell(row=1, column=1, value="Session summary — comparison").font = TITLE_FONT

    def best(rows, lap_idx):
        for r in rows:
            if r["lap"] == lap_idx:
                return r
        v = [r for r in rows if r["valid"] and r["t_ms"]]
        return min(v, key=lambda r: r["t_ms"]) if v else None
    b1, b2 = best(rows1, la), best(rows2, lb)

    headers = ["", n1, n2, "Δ (P1−P2)"]
    hr = 3
    for i, h in enumerate(headers, 1):
        ws.cell(row=hr, column=i, value=h)
    style_header(ws, hr, len(headers))

    def sec(ms):
        return round(ms / 1000, 3) if ms else None
    data = []
    if b1 and b2:
        data = [
            ("Best lap", fmt_ms(b1["t_ms"]), fmt_ms(b2["t_ms"]),
             sec(b1["t_ms"]), sec(b2["t_ms"])),
            ("Sector 1", fmt_ms(b1["s1"]), fmt_ms(b2["s1"]), sec(b1["s1"]), sec(b2["s1"])),
            ("Sector 2", fmt_ms(b1["s2"]), fmt_ms(b2["s2"]), sec(b1["s2"]), sec(b2["s2"])),
            ("Sector 3", fmt_ms(b1["s3"]), fmt_ms(b2["s3"]), sec(b1["s3"]), sec(b2["s3"])),
            ("Top speed", f'{round(b1["top"])} km/h', f'{round(b2["top"])} km/h',
             b1["top"], b2["top"]),
        ]
    for j, (label, v1, v2, n_a, n_b) in enumerate(data):
        rr = hr + 1 + j
        ws.cell(row=rr, column=1, value=label).font = Font(name=FONT, bold=True, size=10)
        ws.cell(row=rr, column=2, value=v1).font = BASE_FONT
        ws.cell(row=rr, column=3, value=v2).font = BASE_FONT
        # Delta formula uses hidden numeric values in columns F and G.
        ws.cell(row=rr, column=6, value=n_a)
        ws.cell(row=rr, column=7, value=n_b)
        dc = ws.cell(row=rr, column=4, value=f"=F{rr}-G{rr}")
        dc.font = BASE_FONT
        dc.number_format = "+0.000;-0.000;0.000"
        for i in range(1, 5):
            ws.cell(row=rr, column=i).border = BORDER
            ws.cell(row=rr, column=i).alignment = Alignment(horizontal="center")
        ws.cell(row=rr, column=1).alignment = Alignment(horizontal="left")
    # Hide the helper numeric columns F and G.
    ws.column_dimensions["F"].hidden = True
    ws.column_dimensions["G"].hidden = True
    for col, w in zip("ABCD", (18, 14, 14, 12)):
        ws.column_dimensions[col].width = w
    note = hr + len(data) + 2
    ws.cell(row=note, column=1,
            value="Negative Δ means P1 is faster. The fastest lap is highlighted "
                  "in green on the lap sheets.").font = Font(name=FONT, italic=True,
                                                            size=9, color="666666")


def write_compare(ws, laps1, laps2, n1, n2):
    ws.cell(row=1, column=1, value="Lap-by-lap comparison").font = TITLE_FONT
    headers = ["Lap", f"{n1} time", f"{n2} time",
               f"{n1} (s)", f"{n2} (s)", "Δ (s)"]
    hr = 3
    for i, h in enumerate(headers, 1):
        ws.cell(row=hr, column=i, value=h)
    style_header(ws, hr, len(headers))

    common = sorted(set(L for L in laps1 if laps1[L]["t"]) &
                    set(L for L in laps2 if laps2[L]["t"]))
    for j, L in enumerate(common):
        rr = hr + 1 + j
        t1, t2 = laps1[L]["t"], laps2[L]["t"]
        ws.cell(row=rr, column=1, value=L)
        ws.cell(row=rr, column=2, value=fmt_ms(t1))
        ws.cell(row=rr, column=3, value=fmt_ms(t2))
        ws.cell(row=rr, column=4, value=round(t1 / 1000, 3))
        ws.cell(row=rr, column=5, value=round(t2 / 1000, 3))
        dc = ws.cell(row=rr, column=6, value=f"=D{rr}-E{rr}")
        dc.number_format = "+0.000;-0.000;0.000"
        for i in range(1, 7):
            c = ws.cell(row=rr, column=i)
            c.font = BASE_FONT
            c.border = BORDER
            c.alignment = Alignment(horizontal="center")
    last = hr + len(common)
    if common:
        ws.conditional_formatting.add(f"F{hr+1}:F{last}", ColorScaleRule(
            start_type="num", start_value=-1, start_color="63BE7B",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="num", end_value=1, end_color="F8696B"))
    for col, w in zip("ABCDEF", (6, 12, 12, 10, 10, 10)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = f"A{hr+1}"


def write_traces(wb, laps1, laps2, la, lb, n1, n2):
    ws = wb.create_sheet("Traces")
    d1, t1 = clean_curve(laps1[la]["samples"], "t")
    d2, t2 = clean_curve(laps2[lb]["samples"], "t")
    _, sp1 = clean_curve(laps1[la]["samples"], "speed")
    _, sp2 = clean_curve(laps2[lb]["samples"], "speed")
    if len(d1) < 2 or len(d2) < 2:
        return
    dmax = min(d1.max(), d2.max())
    grid = np.linspace(0, dmax, 400)
    S1 = np.interp(grid, d1, sp1)
    S2 = np.interp(grid, d2, sp2)
    delta = (np.interp(grid, d1, t1) - np.interp(grid, d2, t2)) / 1000.0

    heads = ["Distance (m)", f"{n1} speed", f"{n2} speed", "Δ time (s)"]
    for i, h in enumerate(heads, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
    for j in range(len(grid)):
        ws.cell(row=2 + j, column=1, value=round(float(grid[j]), 1))
        ws.cell(row=2 + j, column=2, value=round(float(S1[j]), 1))
        ws.cell(row=2 + j, column=3, value=round(float(S2[j]), 1))
        ws.cell(row=2 + j, column=4, value=round(float(delta[j]), 3))
    n = len(grid)

    # Two-series speed chart.
    ch = LineChart()
    ch.title = f"Speed over distance — {n1} lap {la} vs {n2} lap {lb}"
    ch.y_axis.title = "km/h"
    ch.x_axis.title = "Distance (m)"
    ch.height, ch.width = 9, 26
    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=1 + n)
    cats = Reference(ws, min_col=1, min_row=2, max_row=1 + n)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    for s in ch.series:
        s.smooth = False
    ws.add_chart(ch, "F2")

    # Time-delta chart.
    ch2 = LineChart()
    ch2.title = "Δ time over distance (>0 → P1 slower)"
    ch2.y_axis.title = "seconds"
    ch2.x_axis.title = "Distance (m)"
    ch2.height, ch2.width = 9, 26
    d = Reference(ws, min_col=4, max_col=4, min_row=1, max_row=1 + n)
    ch2.add_data(d, titles_from_data=True)
    ch2.set_categories(cats)
    ws.add_chart(ch2, "F20")

    for col, w in zip("ABCD", (13, 12, 12, 11)):
        ws.column_dimensions[col].width = w


# ----------------------------------------------------------------------------
def build(path_a, path_b, out):
    sa, sb = load_rich(path_a), load_rich(path_b)
    rows1, laps1 = lap_table(sa)
    rows2, laps2 = lap_table(sb)
    la = pick_fastest(laps1)
    lb = pick_fastest(laps2)
    if la is None or lb is None:
        sys.exit("No complete laps found in the CSV files (possibly only out-laps).")
    n1, n2 = "P1", "P2"

    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    write_summary(ws_sum, rows1, rows2, la, lb, n1, n2)
    write_lap_sheet(wb.create_sheet("P1 Laps"), f"Laps — {n1}", rows1, P1_FILL)
    write_lap_sheet(wb.create_sheet("P2 Laps"), f"Laps — {n2}", rows2, P2_FILL)
    write_compare(wb.create_sheet("Comparison"), laps1, laps2, n1, n2)
    write_traces(wb, laps1, laps2, la, lb, n1, n2)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    wb.save(out)
    print(f"Report saved: {out}")
    print(f"  Fastest: {n1} lap {la} = {fmt_ms(laps1[la]['t'])}  |  "
          f"{n2} lap {lb} = {fmt_ms(laps2[lb]['t'])}")


def main():
    ap = argparse.ArgumentParser(description="F1 25/26 lap-by-lap Excel report")
    ap.add_argument("csv_p1")
    ap.add_argument("csv_p2")
    ap.add_argument("--out", default="f1_report.xlsx")
    a = ap.parse_args()
    build(a.csv_p1, a.csv_p2, a.out)


if __name__ == "__main__":
    main()
